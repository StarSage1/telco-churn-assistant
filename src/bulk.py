from __future__ import annotations

from copy import copy
from io import BytesIO
import json
import os
import re
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import httpx
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from .inference import predict_churn_batch
    from .reports import create_bulk_report
    from .schemas import CustomerInput
except ImportError:
    from inference import predict_churn_batch
    from reports import create_bulk_report
    from schemas import CustomerInput


MAX_UPLOAD_BYTES = 15 * 1024 * 1024
MAX_ROWS = 20_000


def _key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


COLUMN_ALIASES = {
    **{_key(name): name for name in CustomerInput.model_fields},
    "sex": "gender",
    "senior": "Senior_Citizen",
    "seniorcitizen": "Senior_Citizen",
    "married": "Is_Married",
    "partner": "Is_Married",
    "months": "tenure",
    "tenuremonths": "tenure",
    "phoneservice": "Phone_Service",
    "multiplelines": "Dual",
    "multiplephonelines": "Dual",
    "internet": "Internet_Service",
    "internetservice": "Internet_Service",
    "onlinesecurity": "Online_Security",
    "onlinebackup": "Online_Backup",
    "deviceprotection": "Device_Protection",
    "techsupport": "Tech_Support",
    "technicalsupport": "Tech_Support",
    "streamingtv": "Streaming_TV",
    "streamingmovies": "Streaming_Movies",
    "paperlessbilling": "Paperless_Billing",
    "payment": "Payment_Method",
    "paymentmethod": "Payment_Method",
    "monthlycharges": "Monthly_Charges",
    "monthlycharge": "Monthly_Charges",
    "totalcharges": "Total_Charges",
    "totalcharge": "Total_Charges",
}


def resolve_columns(
    columns: list[str], assisted_mapping: dict[str, str] | None = None
) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for column in columns:
        target = COLUMN_ALIASES.get(_key(column))
        if target and target not in mapping.values():
            mapping[column] = target
    for column, target in (assisted_mapping or {}).items():
        if column in columns and target in CustomerInput.model_fields and target not in mapping.values():
            mapping[column] = target
    missing = [name for name in CustomerInput.model_fields if name not in mapping.values()]
    if missing:
        raise ValueError(
            "The file is missing required customer columns: " + ", ".join(missing)
        )
    return mapping


async def infer_unknown_columns(columns: list[str]) -> dict[str, str]:
    """Ask the local LLM once to map unfamiliar spreadsheet headers."""
    expected = list(CustomerInput.model_fields)
    payload = {
        "model": os.getenv("OLLAMA_MODEL", "qwen3:1.7b"),
        "messages": [
            {
                "role": "system",
                "content": (
                    "Map uploaded telecom spreadsheet headers to the provided model fields. "
                    "Return only a JSON object whose keys are uploaded headers and values are "
                    "exact model field names. Omit unrelated columns. Never invent a header."
                ),
            },
            {
                "role": "user",
                "content": json.dumps({"uploaded_headers": columns, "model_fields": expected}),
            },
        ],
        "format": "json",
        "stream": False,
        "think": False,
        "options": {"temperature": 0},
    }
    try:
        async with httpx.AsyncClient(timeout=45) as client:
            response = await client.post(
                f'{os.getenv("OLLAMA_URL", "http://127.0.0.1:11434")}/api/chat',
                json=payload,
            )
            response.raise_for_status()
        candidate = json.loads(response.json()["message"]["content"])
        if not isinstance(candidate, dict):
            return {}
        return {
            str(column): str(target)
            for column, target in candidate.items()
            if column in columns and target in expected
        }
    except (httpx.HTTPError, KeyError, TypeError, ValueError, json.JSONDecodeError):
        return {}


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _normalize_value(field: str, value: Any) -> Any:
    if pd.isna(value) or (isinstance(value, str) and not value.strip()):
        return None
    if field == "tenure":
        return int(round(float(value)))
    if field in {"Monthly_Charges", "Total_Charges"}:
        return float(str(value).replace(",", "").replace("$", "").strip())

    text = _text(value)
    if field == "Senior_Citizen":
        if text in {"1", "yes", "true", "senior"}:
            return 1
        if text in {"0", "no", "false", "not senior"}:
            return 0

    yes_no_fields = {
        "Is_Married", "Dependents", "Phone_Service", "Paperless_Billing"
    }
    if field in yes_no_fields:
        if text in {"yes", "y", "true", "1"}:
            return "Yes"
        if text in {"no", "n", "false", "0"}:
            return "No"

    aliases = {
        "gender": {"female": "Female", "f": "Female", "male": "Male", "m": "Male"},
        "Dual": {
            "yes": "Yes", "no": "No", "multiple": "Yes", "single": "No",
            "no phone service": "No phone service", "none": "No phone service",
        },
        "Internet_Service": {
            "fiber": "Fiber optic", "fiber optic": "Fiber optic",
            "fiber optics": "Fiber optic", "fibre optic": "Fiber optic",
            "dsl": "DSL", "no": "No", "none": "No", "no internet": "No",
            "no internet service": "No",
        },
        "Contract": {
            "month": "Month-to-month", "monthly": "Month-to-month",
            "month-to-month": "Month-to-month", "month to month": "Month-to-month",
            "one year": "One year", "1 year": "One year",
            "two year": "Two year", "2 year": "Two year",
        },
        "Payment_Method": {
            "electronic": "Electronic check", "electronic check": "Electronic check",
            "e-check": "Electronic check", "mailed check": "Mailed check",
            "mail": "Mailed check", "bank transfer": "Bank transfer (automatic)",
            "bank transfer (automatic)": "Bank transfer (automatic)",
            "automatic bank transfer": "Bank transfer (automatic)",
            "credit card": "Credit card (automatic)",
            "credit card (automatic)": "Credit card (automatic)",
            "automatic credit card": "Credit card (automatic)",
        },
    }
    internet_features = {
        "Online_Security", "Online_Backup", "Device_Protection", "Tech_Support",
        "Streaming_TV", "Streaming_Movies",
    }
    if field in internet_features:
        if text in {"yes", "y", "true", "1"}:
            return "Yes"
        if text in {"no", "n", "false", "0"}:
            return "No"
        if text in {"none", "no internet", "no internet service"}:
            return "No internet service"
    return aliases.get(field, {}).get(text, value)


def read_upload(filename: str, content: bytes) -> pd.DataFrame:
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("File is larger than the 15 MB upload limit.")
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    try:
        if suffix == "csv":
            frame = pd.read_csv(BytesIO(content))
        elif suffix == "xlsx":
            frame = pd.read_excel(BytesIO(content), engine="openpyxl")
        else:
            raise ValueError("Upload a .csv or .xlsx file.")
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("The uploaded spreadsheet could not be read.") from exc
    if frame.empty:
        raise ValueError("The uploaded file has no customer rows.")
    if len(frame) > MAX_ROWS:
        raise ValueError(f"The file exceeds the {MAX_ROWS:,}-row limit.")
    return frame


def score_frame(
    frame: pd.DataFrame, assisted_mapping: dict[str, str] | None = None
) -> tuple[pd.DataFrame, dict[str, Any], list[dict[str, Any]]]:
    column_mapping = resolve_columns(
        [str(column) for column in frame.columns], assisted_mapping
    )
    valid_rows: list[dict[str, Any]] = []
    valid_indexes: list[Any] = []
    errors: dict[Any, str] = {}

    for index, source_row in frame.iterrows():
        try:
            candidate = {
                target: _normalize_value(target, source_row[source])
                for source, target in column_mapping.items()
            }
            customer = CustomerInput.model_validate(candidate)
            valid_rows.append(customer.model_dump())
            valid_indexes.append(index)
        except Exception as exc:
            detail = exc.errors()[0]["msg"] if hasattr(exc, "errors") else str(exc)
            errors[index] = str(detail)

    scored = frame.copy()
    scored["Churn"] = "Error"
    scored["Churn_Percentage"] = pd.NA
    row_results: dict[Any, tuple[dict[str, Any], dict[str, Any]]] = {}

    results = predict_churn_batch(valid_rows)
    for index, customer, result in zip(valid_indexes, valid_rows, results):
        probability = result["churn_probability"]
        scored.at[index, "Churn"] = result["prediction_label"]
        scored.at[index, "Churn_Percentage"] = probability
        row_results[index] = (customer, result)

    valid_count = len(valid_rows)
    churn_count = sum(result["prediction"] for result in results)
    summary = {
        "total_rows": len(scored),
        "scored_rows": valid_count,
        "invalid_rows": len(errors),
        "predicted_churn": churn_count,
        "churn_rate": churn_count / valid_count if valid_count else 0.0,
    }
    id_columns = [
        column for column in frame.columns
        if _key(column) in {"customerid", "customer", "accountid", "id"}
    ]
    findings: list[dict[str, Any]] = []
    for position, index in enumerate(frame.index, 2):
        identifier = (
            str(frame.at[index, id_columns[0]])
            if id_columns and not pd.isna(frame.at[index, id_columns[0]])
            else f"Row {position}"
        )
        if index in row_results:
            customer, result = row_results[index]
            findings.append(
                {
                    "row_number": position,
                    "customer_id": identifier,
                    "customer": customer,
                    "result": result,
                    "error": None,
                }
            )
        else:
            findings.append(
                {
                    "row_number": position,
                    "customer_id": identifier,
                    "customer": None,
                    "result": None,
                    "error": errors.get(index, "The row could not be validated."),
                }
            )
    return scored, summary, findings


def _style_prediction_columns(sheet, first_column: int, row_count: int) -> None:
    navy = "102630"
    white = "FFFFFF"
    header_source = sheet.cell(1, max(first_column - 1, 1))
    for column in (first_column, first_column + 1):
        target = sheet.cell(1, column)
        if header_source.has_style:
            target._style = copy(header_source._style)
        else:
            target.fill = PatternFill("solid", fgColor=navy)
            target.font = Font(color=white, bold=True)
            target.alignment = Alignment(wrap_text=True, vertical="center")
    for row in range(2, row_count + 2):
        source = sheet.cell(row, max(first_column - 1, 1))
        for column in (first_column, first_column + 1):
            if source.has_style:
                sheet.cell(row, column)._style = copy(source._style)
        sheet.cell(row, first_column + 1).number_format = "0.0%"
        churn_cell = sheet.cell(row, first_column)
        if churn_cell.value == "Churn":
            churn_cell.fill = PatternFill("solid", fgColor="FCE4DF")
        elif churn_cell.value == "No Churn":
            churn_cell.fill = PatternFill("solid", fgColor="DDF7F1")
    probability_letter = get_column_letter(first_column + 1)
    if row_count:
        sheet.conditional_formatting.add(
            f"{probability_letter}2:{probability_letter}{row_count + 1}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="DDF7F1",
                mid_type="num", mid_value=0.31, mid_color="FFF1C9",
                end_type="num", end_value=1, end_color="F6B7AA",
            ),
        )
    sheet.column_dimensions[get_column_letter(first_column)].width = 16
    sheet.column_dimensions[get_column_letter(first_column + 1)].width = 20


def create_cloned_workbook(
    filename: str, content: bytes, scored: pd.DataFrame
) -> bytes:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    original_column_count = len(scored.columns) - 2
    if suffix == "xlsx":
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.worksheets[0]
        first_prediction_column = original_column_count + 1
        sheet.cell(1, first_prediction_column).value = "Churn"
        sheet.cell(1, first_prediction_column + 1).value = "Churn_Percentage"
        for output_row, (_, row) in enumerate(scored.iterrows(), 2):
            sheet.cell(output_row, first_prediction_column).value = row["Churn"]
            probability = row["Churn_Percentage"]
            sheet.cell(output_row, first_prediction_column + 1).value = (
                None if pd.isna(probability) else float(probability)
            )
        _style_prediction_columns(sheet, first_prediction_column, len(scored))
        for table in sheet.tables.values():
            from openpyxl.utils.cell import range_boundaries

            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if min_row == 1 and max_col == original_column_count:
                table.ref = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(first_prediction_column + 1)}{max_row}"
                )
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"
        safe_frame = scored.astype(object).where(pd.notna(scored), None)
        sheet.append([str(column) for column in safe_frame.columns])
        for row in safe_frame.itertuples(index=False, name=None):
            sheet.append(list(row))
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        first_prediction_column = original_column_count + 1
        _style_prediction_columns(sheet, first_prediction_column, len(scored))
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="102630")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        table = Table(displayName="CustomersWithChurn", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
        for column_index, column in enumerate(safe_frame.columns, 1):
            width = min(max(len(str(column)) + 2, 12), 30)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def process_bulk_file_assisted(
    filename: str, content: bytes
) -> tuple[bytes, dict[str, Any]]:
    """Parse known schemas immediately and use Ollama only for unfamiliar headers."""
    frame = read_upload(filename, content)
    try:
        scored, summary, findings = score_frame(frame)
    except ValueError as original_error:
        assisted = await infer_unknown_columns([str(column) for column in frame.columns])
        try:
            scored, summary, findings = score_frame(frame, assisted)
        except ValueError:
            raise original_error
    workbook = create_cloned_workbook(filename, content, scored)
    report = create_bulk_report(findings, summary, filename)
    safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "-", filename.rsplit(".", 1)[0]).strip("-")
    safe_stem = safe_stem or "customers"
    archive = BytesIO()
    with ZipFile(archive, "w", compression=ZIP_DEFLATED) as package:
        package.writestr(f"{safe_stem}-with-churn.xlsx", workbook)
        package.writestr(f"{safe_stem}-findings.pdf", report)
    return archive.getvalue(), summary
