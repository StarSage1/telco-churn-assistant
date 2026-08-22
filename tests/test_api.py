import csv
from io import BytesIO, StringIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

from src.Api import app
from tests.test_chatbot import COMPLETE_CUSTOMER


client = TestClient(app)


def test_health_endpoint():
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json()["service"] == "telco-churn-assistant"


def test_prediction_endpoint_uses_saved_models():
    response = client.post("/predict", json=COMPLETE_CUSTOMER)
    assert response.status_code == 200
    body = response.json()
    assert body["prediction_label"] in {"Churn", "No Churn"}
    assert 0 <= body["churn_probability"] <= 1


def test_prediction_accepts_monthly_charge_above_previous_artificial_cap():
    customer = {**COMPLETE_CUSTOMER, "Monthly_Charges": 714.5}
    response = client.post("/predict", json=customer)

    assert response.status_code == 200
    assert 0 <= response.json()["churn_probability"] <= 1


def test_prediction_rejects_logically_conflicting_services():
    customer = {**COMPLETE_CUSTOMER, "Phone_Service": "No", "Dual": "Yes"}
    response = client.post("/predict", json=customer)
    assert response.status_code == 422


def test_professional_customer_report_is_a_readable_pdf():
    response = client.post("/report", json=COMPLETE_CUSTOMER)

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    reader = PdfReader(BytesIO(response.content))
    assert len(reader.pages) == 1
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Customer Churn Assessment" in text
    assert "NEXT BEST ACTION" in text
    assert "CUSTOMER PROFILE" in text


def test_bulk_csv_returns_excel_clone_and_row_findings_pdf():
    source = StringIO()
    writer = csv.DictWriter(source, fieldnames=list(COMPLETE_CUSTOMER))
    writer.writeheader()
    writer.writerow(COMPLETE_CUSTOMER)
    writer.writerow({**COMPLETE_CUSTOMER, "tenure": 60, "Contract": "Two year"})

    response = client.post(
        "/bulk-predict",
        files={"file": ("customers.csv", source.getvalue(), "text/csv")},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    assert response.headers["x-total-rows"] == "2"
    assert response.headers["x-scored-rows"] == "2"
    package = ZipFile(BytesIO(response.content))
    assert set(package.namelist()) == {
        "customers-with-churn.xlsx",
        "customers-findings.pdf",
    }
    workbook = load_workbook(BytesIO(package.read("customers-with-churn.xlsx")))
    sheet = workbook["Customers"]
    headers = [cell.value for cell in sheet[1]]
    assert headers[-2:] == ["Churn", "Churn_Percentage"]
    assert sheet.max_row == 3
    report = PdfReader(BytesIO(package.read("customers-findings.pdf")))
    text = "\n".join(page.extract_text() or "" for page in report.pages)
    assert "Customer Findings and Actions" in text
    assert "Row 2" in text
    assert "Row 3" in text


def test_bulk_csv_marks_malformed_numeric_row_invalid_without_aborting_file():
    source = StringIO()
    writer = csv.DictWriter(source, fieldnames=list(COMPLETE_CUSTOMER))
    writer.writeheader()
    writer.writerow(COMPLETE_CUSTOMER)
    writer.writerow({**COMPLETE_CUSTOMER, "Monthly_Charges": "not-a-number"})

    response = client.post(
        "/bulk-predict",
        files={"file": ("mixed-validity.csv", source.getvalue(), "text/csv")},
    )

    assert response.status_code == 200
    assert response.headers["x-total-rows"] == "2"
    assert response.headers["x-scored-rows"] == "1"
    assert response.headers["x-invalid-rows"] == "1"

    package = ZipFile(BytesIO(response.content))
    workbook = load_workbook(
        BytesIO(package.read("mixed-validity-with-churn.xlsx"))
    )
    sheet = workbook["Customers"]
    assert sheet.cell(2, sheet.max_column - 1).value in {"Churn", "No Churn"}
    assert sheet.cell(3, sheet.max_column - 1).value == "Error"
    assert sheet.cell(3, sheet.max_column).value is None


def test_bulk_xlsx_upload_is_supported():
    source_workbook = Workbook()
    source_sheet = source_workbook.active
    source_sheet.append(list(COMPLETE_CUSTOMER))
    source_sheet.append(list(COMPLETE_CUSTOMER.values()))
    notes = source_workbook.create_sheet("Original Notes")
    notes["A1"] = "This sheet must remain unchanged"
    source = BytesIO()
    source_workbook.save(source)

    response = client.post(
        "/bulk-predict",
        files={
            "file": (
                "customers.xlsx",
                source.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.headers["x-total-rows"] == "1"
    assert response.headers["x-scored-rows"] == "1"
    package = ZipFile(BytesIO(response.content))
    cloned = load_workbook(BytesIO(package.read("customers-with-churn.xlsx")))
    assert cloned.sheetnames == ["Sheet", "Original Notes"]
    assert cloned["Original Notes"]["A1"].value == "This sheet must remain unchanged"
    assert [cell.value for cell in cloned["Sheet"][1]][-2:] == [
        "Churn",
        "Churn_Percentage",
    ]
